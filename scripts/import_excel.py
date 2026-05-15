"""
3개 엑셀 파일을 통합하여 ulsan_schools.json을 생성하는 스크립트.

소스:
  1. 202603학교학생수학급수현황.xlsx (5개 시트: 초/중/고/특수/각종)
       → 학년별 학급수·학생수
  2. 울산학교주소위도경도.xlsx → 주소, 위도, 경도 (위도/경도는 보정된 값)
  3. 울산학교개교일우편번호전화번호팩스번호홈페이지.xlsx
       → 개교일, 우편번호, 전화, 팩스, 홈페이지
"""

import json
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).parent.parent
XLSX_DIR = BASE_DIR / "data" / "xlsx"
STATUS_XLSX = XLSX_DIR / "202603학교학생수학급수현황.xlsx"
LOC_XLSX = XLSX_DIR / "울산학교주소위도경도.xlsx"
CONTACT_XLSX = XLSX_DIR / "울산학교개교일우편번호전화번호팩스번호홈페이지.xlsx"
DATA_PATH = BASE_DIR / "data" / "ulsan_schools.json"
# 학구 GeoJSON (convert_school_zones.py가 생성). 각 feature.properties 의
# elementary_schools / middle_schools / high_schools 리스트를 역참조해
# 학교별 elem_zone / middle_zone / high_zone 필드를 다시 채운다.
ZONES_ELEM = BASE_DIR / "data" / "ulsan_school_zones.json"
ZONES_MID = BASE_DIR / "data" / "ulsan_middle_school_zones.json"
ZONES_HIGH = BASE_DIR / "data" / "ulsan_high_school_zones.json"


# ── 1) 현황 엑셀 읽기 (단일 파일 · 5개 시트) ─────────────────────

def _to_int(v):
    return int(v) if v not in (None, "") else 0


def _strip(v):
    return str(v).strip() if v not in (None, "") else ""


def read_elementary(ws):
    """[초] 시트 → 초등학교 리스트.

    컬럼: 교육지원청, 설립구분, 학교, 자치구,
          학급수_1~6학년, 학생수_1~6학년_계
    """
    schools = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = _strip(row[2])
        if not name:
            continue
        cbg = {str(g): _to_int(row[4 + g - 1]) for g in range(1, 7)}
        sbg = {str(g): _to_int(row[10 + g - 1]) for g in range(1, 7)}
        schools.append({
            "name": name,
            "type": "초등학교",
            "founding_type": _strip(row[1]),
            "district": _strip(row[3]),
            "total_classes": sum(cbg.values()),
            "classes_by_grade": cbg,
            "total_students": sum(sbg.values()),
            "students_by_grade": sbg,
        })
    return schools


def read_middle(ws):
    """[중] 시트 → 중학교 리스트.

    컬럼: 교육지원청, 설립구분, 학교, 자치구, 주야구분, 남녀공학구분,
          학급수_1~3학년, 학생수_1~3학년_계
    """
    schools = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = _strip(row[2])
        if not name:
            continue
        cbg = {str(g): _to_int(row[6 + g - 1]) for g in range(1, 4)}
        sbg = {str(g): _to_int(row[9 + g - 1]) for g in range(1, 4)}
        schools.append({
            "name": name,
            "type": "중학교",
            "founding_type": _strip(row[1]),
            "district": _strip(row[3]),
            "coedu": _strip(row[5]),
            "total_classes": sum(cbg.values()),
            "classes_by_grade": cbg,
            "total_students": sum(sbg.values()),
            "students_by_grade": sbg,
        })
    return schools


def read_high(ws):
    """[고] 시트 → 고등학교 리스트.

    컬럼: 교육지원청, 학교급, 설립구분, 학교, 자치구,
          고교유형, 주야구분, 남녀공학구분,
          학급수_계_1~3학년, 학생수_1~3학년_계
    """
    schools = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = _strip(row[3])
        if not name:
            continue
        cbg = {str(g): _to_int(row[8 + g - 1]) for g in range(1, 4)}
        sbg = {str(g): _to_int(row[11 + g - 1]) for g in range(1, 4)}
        schools.append({
            "name": name,
            "type": "고등학교",
            "founding_type": _strip(row[2]),
            "district": _strip(row[4]),
            "hs_category": _strip(row[5]),
            "coedu": _strip(row[7]),
            "total_classes": sum(cbg.values()),
            "classes_by_grade": cbg,
            "total_students": sum(sbg.values()),
            "students_by_grade": sbg,
        })
    return schools


def read_special(ws):
    """[특수] 시트 → 특수학교 리스트.

    컬럼: 교육지원청, 설립구분, 조직(학교명), 자치구,
          학급수_유치원/초등학교/중학교/고등학교/전공과,
          학생수_유치원/초등학교/중학교/고등학교/전공과_계
    학년 키는 UI 호환을 위해 "유치원/초등/중/고/전공과" 로 유지.
    """
    stages = ["유치원", "초등", "중", "고", "전공과"]
    schools = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = _strip(row[2])
        if not name:
            continue
        cbg = {stages[i]: _to_int(row[4 + i]) for i in range(5)}
        sbg = {stages[i]: _to_int(row[9 + i]) for i in range(5)}
        schools.append({
            "name": name,
            "type": "특수학교",
            "founding_type": _strip(row[1]),
            "district": _strip(row[3]),
            "total_classes": sum(cbg.values()),
            "classes_by_grade": cbg,
            "total_students": sum(sbg.values()),
            "students_by_grade": sbg,
        })
    return schools


def read_alternative(ws):
    """[각종] 시트 → 각종학교 리스트.

    컬럼: 교육지원청, 설립구분, 조직(학교명), 자치구,
          1학년학급수, 1학년학생수, 2학년학급수, 2학년학생수, 3학년학급수, 3학년학생수
    (각종학교 시트는 학급수와 학생수가 교차 배치됨)
    """
    schools = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = _strip(row[2])
        if not name:
            continue
        cbg = {str(g): _to_int(row[4 + (g - 1) * 2]) for g in range(1, 4)}
        sbg = {str(g): _to_int(row[5 + (g - 1) * 2]) for g in range(1, 4)}
        schools.append({
            "name": name,
            "type": "각종학교",
            "founding_type": _strip(row[1]),
            "district": _strip(row[3]),
            "total_classes": sum(cbg.values()),
            "classes_by_grade": cbg,
            "total_students": sum(sbg.values()),
            "students_by_grade": sbg,
        })
    return schools


def read_status(path):
    """현황 단일 워크북에서 5개 시트를 읽어 학교 리스트를 합쳐 반환."""
    wb = openpyxl.load_workbook(path, data_only=True)
    elem = read_elementary(wb["초"])
    mid = read_middle(wb["중"])
    high = read_high(wb["고"])
    special = read_special(wb["특수"])
    alt = read_alternative(wb["각종"])
    return elem, mid, high, special, alt


# ── 2) 위치 엑셀 읽기 ──────────────────────────────────────────

def read_location(path):
    """울산학교주소위도경도.xlsx → {학교명: {address, lat, lng}}"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    loc = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = _strip(row[0])
        if not name:
            continue
        addr_val = row[7] if len(row) > 7 else ""
        lat_val = row[8] if len(row) > 8 else 0
        lng_val = row[9] if len(row) > 9 else 0
        loc[name] = {
            "address": _strip(addr_val),
            "lat": float(lat_val) if lat_val not in (None, "") else 0,
            "lng": float(lng_val) if lng_val not in (None, "") else 0,
        }
    return loc


# ── 3) 연락처 엑셀 읽기 ────────────────────────────────────────

def normalize_homepage(url):
    if not url or url.strip() in ("", "http://", "https://"):
        return ""
    url = url.strip()
    if not url.startswith("http://") and not url.startswith("https://"):
        url = "http://" + url
    return url


def read_contact(path):
    """울산학교개교일...홈페이지.xlsx → {학교명: {founded, zipcode, phone, fax, homepage}}"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    info = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = _strip(row[1])
        if not name:
            continue
        info[name] = {
            "founded": _strip(row[2]),
            "zipcode": _strip(row[3]),
            "phone": _strip(row[4]),
            "fax": _strip(row[5]),
            "homepage": normalize_homepage(str(row[6] or "")),
        }
    return info


# ── 4) 이름 매칭 헬퍼 ──────────────────────────────────────────

def find_match(name, lookup):
    """학교명 매칭: 직접 → 울산 접두사 추가 → 울산 제거"""
    if name in lookup:
        return lookup[name]
    if ("울산" + name) in lookup:
        return lookup["울산" + name]
    short = name.replace("울산", "", 1)
    if short != name and short in lookup:
        return lookup[short]
    return None


def matched_key(name, lookup):
    """find_match와 동일한 규칙이지만 매칭된 lookup 키를 반환."""
    if name in lookup:
        return name
    if ("울산" + name) in lookup:
        return "울산" + name
    short = name.replace("울산", "", 1)
    if short != name and short in lookup:
        return short
    return None


# ── 4b) 학구 필드 채우기 (캐시된 GeoJSON 역참조) ─────────────────

def _zone_lookup(geojson_path, school_list_key):
    """학구 GeoJSON → {학교명: 학구이름} 매핑."""
    if not geojson_path.exists():
        return {}
    with open(geojson_path, encoding="utf-8") as f:
        gj = json.load(f)
    mapping = {}
    for feat in gj.get("features", []):
        props = feat.get("properties", {})
        zone_name = props.get("name")
        for sn in props.get(school_list_key, []) or []:
            mapping[sn] = zone_name
    return mapping


def apply_zone_fields(all_schools):
    """학교 dict 에 elem_zone / middle_zone / high_zone 필드 채움.

    convert_school_zones.py가 미리 생성한 GeoJSON 안에 학교 리스트가
    캐시되어 있으므로 SHP/공간 라이브러리 없이도 매번 일관되게 복원 가능.
    """
    elem_map = _zone_lookup(ZONES_ELEM, "elementary_schools")
    mid_map = _zone_lookup(ZONES_MID, "middle_schools")
    high_map = _zone_lookup(ZONES_HIGH, "high_schools")

    e_cnt = m_cnt = h_cnt = 0
    for s in all_schools:
        name = s["name"]
        if s["type"] == "초등학교" and name in elem_map:
            s["elem_zone"] = elem_map[name]
            e_cnt += 1
        if s["type"] in ("초등학교", "중학교") and name in mid_map:
            s["middle_zone"] = mid_map[name]
            m_cnt += 1
        if s["type"] == "고등학교" and name in high_map:
            s["high_zone"] = high_map[name]
            h_cnt += 1
    print(f"학구 매핑: elem_zone {e_cnt}, middle_zone {m_cnt}, high_zone {h_cnt}")


# ── 5) 통계 계산 ───────────────────────────────────────────────

def calc_stats(s):
    tc = s["total_classes"]
    ts = s["total_students"]
    s["avg_students_per_class"] = round(ts / tc, 1) if tc > 0 else 0

    avg_bg = {}
    for g, cnt in s["classes_by_grade"].items():
        stu = s["students_by_grade"].get(g, 0)
        avg_bg[g] = round(stu / cnt, 1) if cnt > 0 else 0
    s["avg_students_per_class_by_grade"] = avg_bg


# ── main ────────────────────────────────────────────────────────

def main():
    # 1) 현황 엑셀 → 학교 기본 데이터
    elem, mid, high, special, alt = read_status(STATUS_XLSX)
    all_schools = elem + mid + high + special + alt
    print(
        f"현황 엑셀: {len(all_schools)}개 "
        f"(초{len(elem)} + 중{len(mid)} + 고{len(high)} + 특수{len(special)} + 각종{len(alt)})"
    )

    # 2) 위치 엑셀 → 주소/위도/경도
    loc = read_location(LOC_XLSX)
    print(f"위치 데이터: {len(loc)}개")

    loc_ok, loc_miss = 0, []
    used_loc_keys = set()
    for s in all_schools:
        m = find_match(s["name"], loc)
        if m:
            s["address"] = m["address"]
            s["lat"] = round(m["lat"], 7)
            s["lng"] = round(m["lng"], 7)
            loc_ok += 1
            k = matched_key(s["name"], loc)
            if k:
                used_loc_keys.add(k)
        else:
            s["address"] = ""
            s["lat"] = 0
            s["lng"] = 0
            loc_miss.append(s["name"])
    print(f"  위치 매칭: {loc_ok}/{len(all_schools)}")
    if loc_miss:
        print(f"  [신규 학교 의심] 현황에 있지만 위치 엑셀에 없음: {loc_miss}")

    # 3) 연락처 엑셀 → 개교일/우편번호/전화/팩스/홈페이지
    contact = read_contact(CONTACT_XLSX)
    print(f"연락처 데이터: {len(contact)}개")

    ct_ok, ct_miss = 0, []
    used_ct_keys = set()
    for s in all_schools:
        m = find_match(s["name"], contact)
        if m:
            s.update(m)
            ct_ok += 1
            k = matched_key(s["name"], contact)
            if k:
                used_ct_keys.add(k)
        else:
            s.update({"founded": "", "zipcode": "", "phone": "", "fax": "", "homepage": ""})
            ct_miss.append(s["name"])
    print(f"  연락처 매칭: {ct_ok}/{len(all_schools)}")
    if ct_miss:
        print(f"  [신규 학교 의심] 현황에 있지만 연락처 엑셀에 없음: {ct_miss}")

    # 3b) 폐교/개명 의심: 위치·연락처 엑셀에는 있지만 현황에는 없는 학교
    loc_orphans = sorted(set(loc.keys()) - used_loc_keys)
    ct_orphans = sorted(set(contact.keys()) - used_ct_keys)
    if loc_orphans or ct_orphans:
        print()
        print("[폐교/개명 의심] 위치·연락처 엑셀에는 있지만 현황 엑셀에 없는 학교:")
        if loc_orphans:
            print(f"  위치 엑셀 잔존: {loc_orphans}")
        if ct_orphans:
            print(f"  연락처 엑셀 잔존: {ct_orphans}")

    # 3c) 중복 좌표 감지 및 자동 오프셋
    # 같은 건물·부지에 둘 이상 학교가 있는 경우(예: 울산고운중·고) 지도에서
    # 마커가 완전히 겹쳐 하나만 보이는 문제 방지. xlsx 원본은 건드리지 않고
    # JSON 단계에서만 학교명 가나다순 2번째부터 약 16m(위도 0.00015) 간격 오프셋.
    coord_groups = {}
    for s in all_schools:
        if s["lat"] and s["lng"]:
            key = (round(s["lat"], 6), round(s["lng"], 6))
            coord_groups.setdefault(key, []).append(s)

    dup_adjusted = 0
    for group in coord_groups.values():
        if len(group) <= 1:
            continue
        group.sort(key=lambda x: x["name"])
        for i, s in enumerate(group[1:], start=1):
            s["lat"] = round(s["lat"] + 0.00015 * i, 7)
            dup_adjusted += 1
    if dup_adjusted > 0:
        print(f"중복 좌표 오프셋: {dup_adjusted}건 (같은 부지 학교 마커 분리)")

    # 4) 학구 필드 채우기 (elem_zone/middle_zone/high_zone)
    apply_zone_fields(all_schools)

    # 5) 통계 계산
    for s in all_schools:
        calc_stats(s)

    # 5) 정렬 및 ID 부여
    all_schools.sort(key=lambda x: (x["type"], x["name"]))
    for i, s in enumerate(all_schools, 1):
        s["id"] = i

    # 6) 메타데이터
    types, districts = {}, {}
    for s in all_schools:
        types[s["type"]] = types.get(s["type"], 0) + 1
        districts[s["district"]] = districts.get(s["district"], 0) + 1

    dataset = {
        "metadata": {
            "title": "울산광역시 학교 데이터",
            "source": "학교급별현황 + 학교위치표준데이터 + NEIS 학교정보",
            "coordinate_system": "WGS84 (EPSG:4326)",
            "last_updated": "2026-03",
            "total_count": len(all_schools),
            "types": types,
            "districts": districts,
        },
        "schools": all_schools,
    }

    with open(DATA_PATH, "w", encoding="utf-8") as f:
        json.dump(dataset, f, ensure_ascii=False, indent=2)

    print(f"\n=== 최종 ===")
    print(f"총 {len(all_schools)}개 학교")
    for t, c in sorted(types.items()):
        print(f"  {t}: {c}")
    print(f"저장: {DATA_PATH}")


if __name__ == "__main__":
    main()
