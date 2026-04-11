"""
초등학교별 연도별 입학생 수 예측 파이프라인

입력:
  1. 행정안전부_지역별(행정동) 성별 연령별 주민등록 인구수_20260331.xlsx
     → 행정동 × 1세 단위 인구
  2. data.go.kr API (통반 단위 총인구)  serviceKey 필요
  3. 통반별학구.xlsx — 통반 → 취학학교 매핑

처리:
  A. 행정동 연령별 인구 추출 (울산 55개 동, 만 0~12세)
  B. API로 울산 통반 총인구 수집
  C. 통반별학구 파싱 (반 쪼개짐은 단순화: 같은 통이 여러 학교면 균등 분배)
  D. 통반 만 N세 추정 = (통반 총인구 / 동 총인구) × 동 만 N세
  E. 학교별 연도별 입학 예정자 집계

출력:
  data/ulsan_dong_age.json            — 동별 연령별 인구 + 매핑용
  data/ulsan_tongban_pop.json         — 통반별 총인구
  data/ulsan_enrollment_forecast.json — 학교별 연도별 예상 입학생
"""
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
import openpyxl
from collections import defaultdict

API_KEY = '8ceef6b4df9fbb7a1603cda86733856cb53d83fc1442b86cc968b70153831541'
API_URL = 'https://apis.data.go.kr/1741000/admmPpltnHhStus/selectAdmmPpltnHhStus'
QUERY_YM = '202602'  # 2026년 2월 말 기준 (현재 날짜 기준 가장 가까운 최신)

DONG_XLSX = '행정안전부_지역별(행정동) 성별 연령별 주민등록 인구수_20260331.xlsx'
ZONE_XLSX = '통반별학구.xlsx'

OUT_DONG = 'data/ulsan_dong_age.json'
OUT_TONGBAN = 'data/ulsan_tongban_pop.json'
OUT_FORECAST = 'data/ulsan_enrollment_forecast.json'

# 한국 초등 입학: 만 6세
# 현재 2026년 4월 기준 만 N세 → (2026 + (6 - N))년 입학 예정
#   N=6 → 2026년 입학
#   N=5 → 2027년 입학
#   N=4 → 2028년 입학
#   N=3 → 2029년 입학
#   N=2 → 2030년 입학
#   N=1 → 2031년 입학
#   N=0 → 2032년 입학
FORECAST_YEARS = [2026, 2027, 2028, 2029, 2030, 2031, 2032]


def step_a_parse_dong():
    """행정동 × 1세 연령별 인구 추출 (울산만)"""
    print('\n[A] 행정동 엑셀 파싱...')
    wb = openpyxl.load_workbook(DONG_XLSX, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    # 열 인덱스 찾기
    col_code = header.index('행정기관코드')
    col_ctpv = header.index('시도명')
    col_sgg = header.index('시군구명')
    col_dong = header.index('읍면동명')
    col_total = header.index('계')

    # 0~12세 남/여 열 인덱스
    age_cols = {}
    for age in range(0, 13):
        try:
            age_cols[age] = {
                'male': header.index(f'{age}세남자'),
                'female': header.index(f'{age}세여자'),
            }
        except ValueError:
            age_cols[age] = None

    result = {}
    for r in rows[1:]:
        if not r or r[col_ctpv] != '울산광역시':
            continue
        code = str(r[col_code])
        dong_name = r[col_dong]
        sgg_name = r[col_sgg]
        total = r[col_total] or 0

        ages = {}
        for age, cols in age_cols.items():
            if cols:
                m = r[cols['male']] or 0
                f = r[cols['female']] or 0
                ages[age] = m + f

        result[code] = {
            'admm_cd': code,
            'sgg': sgg_name,
            'dong': dong_name,
            'total': total,
            'ages': ages,  # {0: n, 1: n, ..., 12: n}
        }
    wb.close()

    print(f'  울산 행정동: {len(result)}개')
    os.makedirs('data', exist_ok=True)
    with open(OUT_DONG, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f'  → {OUT_DONG}')
    return result


def step_b_fetch_tongban(dong_map):
    """각 행정동에 대해 API 호출 → 통반 단위 총인구"""
    print('\n[B] 통반 API 수집...')
    tongban_data = []

    for i, (code, dong) in enumerate(dong_map.items()):
        page = 1
        dong_tongban = []
        while True:
            params = {
                'serviceKey': API_KEY,
                'admmCd': code,
                'srchFrYm': QUERY_YM,
                'srchToYm': QUERY_YM,
                'lv': '4',
                'type': 'JSON',
                'numOfRows': '100',
                'pageNo': str(page),
            }
            url = API_URL + '?' + urllib.parse.urlencode(params, safe=':/')
            try:
                with urllib.request.urlopen(url, timeout=15) as resp:
                    raw = resp.read().decode('utf-8')
                    data = json.loads(raw)
            except Exception as e:
                print(f'  [!] {code} page {page} 실패: {e}')
                break

            head = data.get('Response', {}).get('head', {})
            if head.get('resultCode') != '0':
                print(f'  [!] {code} API error: {head.get("resultMsg")}')
                break

            items = data.get('Response', {}).get('items', {}).get('item', [])
            if isinstance(items, dict):
                items = [items]
            for it in items:
                dong_tongban.append({
                    'admm_cd': code,
                    'dong': it.get('dongNm'),
                    'tong': int(it.get('tong') or 0),
                    'ban': int(it.get('ban') or 0),
                    'total': int(it.get('totNmprCnt') or 0),
                    'male': int(it.get('maleNmprCnt') or 0),
                    'female': int(it.get('femlNmprCnt') or 0),
                })

            total_count = int(head.get('totalCount') or 0)
            if page * 100 >= total_count:
                break
            page += 1
            time.sleep(0.1)

        tongban_data.extend(dong_tongban)
        print(f'  [{i+1}/{len(dong_map)}] {dong["dong"]}: {len(dong_tongban)}개 통반')
        time.sleep(0.15)  # API 호출 간격

    with open(OUT_TONGBAN, 'w', encoding='utf-8') as f:
        json.dump(tongban_data, f, ensure_ascii=False, indent=2)
    print(f'  총 {len(tongban_data)}개 통반 → {OUT_TONGBAN}')
    return tongban_data


def step_c_parse_zones():
    """통반별학구 엑셀 파싱 + '통' 단위 정규화

    파일 형식: 구군, 취학학교명, 읍면동, 통
    통 값: "1통", "3통(1-2반)", "3통 5반", "1통 4반(139~145번지)" 등 복잡함

    단순화 정책:
    - 같은 동+통이 여러 학교에 나뉘면 균등 분배 (1/n 씩)
    - 반 세부 분할과 지번 표기는 무시
    """
    print('\n[C] 통반별학구 엑셀 파싱...')
    wb = openpyxl.load_workbook(ZONE_XLSX, read_only=True)
    ws = wb['Sheet1']
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    print(f'  header: {header}')

    # 통 번호만 추출 ("3통(1-2반)" → 3)
    tong_re = re.compile(r'(\d+)\s*통')

    # (dong, tong) → [school, school, ...]
    mapping = defaultdict(set)
    schools = set()
    skipped = 0

    for r in rows[1:]:
        if not r or not r[1]:
            continue
        gu = r[0]
        school = r[1]
        dong = r[2]
        tong_str = r[3] or ''
        m = tong_re.search(str(tong_str))
        if not m:
            skipped += 1
            continue
        tong = int(m.group(1))
        key = (dong, tong)
        mapping[key].add(school)
        schools.add(school)

    wb.close()

    # 통당 학교 수 분포
    multi = sum(1 for v in mapping.values() if len(v) > 1)
    print(f'  매핑: {len(mapping)}개 (동,통) 쌍, 학교 {len(schools)}개')
    print(f'  복수 학교 쌍: {multi}개 (균등 분배 처리)')
    print(f'  파싱 실패 행: {skipped}개')

    # 직렬화 가능한 형식으로 변환
    mapping_out = {}
    for (dong, tong), school_set in mapping.items():
        key = f'{dong}|{tong}'
        mapping_out[key] = sorted(school_set)
    return mapping_out


def step_d_estimate(dong_map, tongban_data, zone_mapping):
    """통반 연령 추정 + 학교별 연도별 입학 예정자 집계"""
    print('\n[D] 추정·집계...')

    # 동별로 통반을 그룹핑, 총합
    by_dong = defaultdict(list)
    for tb in tongban_data:
        by_dong[tb['admm_cd']].append(tb)

    # admm_cd → dong name (엑셀 동 이름과 매칭 위해)
    code_to_dong = {code: d['dong'] for code, d in dong_map.items()}

    # 학교별 연도별 인구 누적
    # forecast[school][year] = 예상 입학생 수
    forecast = defaultdict(lambda: defaultdict(float))
    # 학구 매칭 통계
    matched_tongban = 0
    unmatched_tongban = 0
    multi_school_tongban = 0

    # 각 통반 처리
    for code, dong_tongban_list in by_dong.items():
        dong_info = dong_map.get(code)
        if not dong_info:
            continue
        dong_name = dong_info['dong']
        dong_total = dong_info['total']
        dong_ages = dong_info['ages']
        if dong_total == 0:
            continue

        # 같은 통 내 여러 반 합산 → 통 단위 총인구
        tong_total = defaultdict(int)
        for tb in dong_tongban_list:
            tong_total[tb['tong']] += tb['total']

        # 각 통에 대해 연령 추정 + 학교 매핑
        for tong, tong_pop in tong_total.items():
            # 해당 통의 만 N세 추정
            tong_age = {}
            for age, dong_age_pop in dong_ages.items():
                tong_age[age] = (tong_pop / dong_total) * dong_age_pop

            # 학교 매핑 조회
            key = f'{dong_name}|{tong}'
            schools = zone_mapping.get(key)
            if not schools:
                unmatched_tongban += 1
                continue
            matched_tongban += 1
            if len(schools) > 1:
                multi_school_tongban += 1

            # 균등 분배 (1/n)
            share = 1.0 / len(schools)
            for school in schools:
                # 연도별 집계
                for year in FORECAST_YEARS:
                    # 해당 연도에 입학하는 나이: (6 - (year - 2026)) ... wait, reverse
                    # year = 2026 → 만 6세 / year = 2027 → 만 5세 / ...
                    age = 6 - (year - 2026)
                    if 0 <= age <= 12 and age in tong_age:
                        forecast[school][year] += tong_age[age] * share

    print(f'  통-학구 매칭: {matched_tongban}개 성공, {unmatched_tongban}개 실패')
    print(f'  여러 학교에 걸친 통: {multi_school_tongban}개')

    # 결과 정리
    result = []
    for school in sorted(forecast.keys()):
        years = forecast[school]
        result.append({
            'school': school,
            'forecast': {str(y): round(years.get(y, 0), 1) for y in FORECAST_YEARS},
            'total_7yr': round(sum(years.values()), 1),
        })

    result.sort(key=lambda x: -x['total_7yr'])

    with open(OUT_FORECAST, 'w', encoding='utf-8') as f:
        json.dump({
            'base_ym': QUERY_YM,
            'forecast_years': FORECAST_YEARS,
            'schools': result,
        }, f, ensure_ascii=False, indent=2)
    print(f'  학교 {len(result)}개 → {OUT_FORECAST}')
    print(f'\nTOP 10 (7년 누적):')
    for r in result[:10]:
        yrs = ', '.join(f'{y}={r["forecast"][str(y)]:.0f}' for y in FORECAST_YEARS)
        print(f'  {r["school"]}: {r["total_7yr"]:.0f}명 ({yrs})')


def main():
    dong_map = step_a_parse_dong()
    tongban = step_b_fetch_tongban(dong_map)
    zones = step_c_parse_zones()
    step_d_estimate(dong_map, tongban, zones)
    print('\n완료.')


if __name__ == '__main__':
    main()
