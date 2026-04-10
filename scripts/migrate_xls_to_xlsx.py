"""
일회성 마이그레이션: 울산학교주소위도경도.xls → .xlsx

- 원본 .xls를 xlrd로 읽고
- data/ulsan_schools.json의 보정된 위도/경도로 덮어쓴 뒤
- openpyxl로 .xlsx로 저장한다.

이후에는 scripts/import_excel.py가 openpyxl로 .xlsx를 직접 읽으므로
xlrd 의존성을 제거할 수 있다.
"""

import json
from pathlib import Path

import xlrd
from openpyxl import Workbook

BASE_DIR = Path(__file__).parent.parent
SRC_XLS = BASE_DIR / "울산학교주소위도경도.xls"
DST_XLSX = BASE_DIR / "울산학교주소위도경도.xlsx"
JSON_PATH = BASE_DIR / "data" / "ulsan_schools.json"


def find_match(name, lookup):
    """import_excel.py의 find_match와 동일한 규칙."""
    if name in lookup:
        return lookup[name]
    if ("울산" + name) in lookup:
        return lookup["울산" + name]
    short = name.replace("울산", "", 1)
    if short != name and short in lookup:
        return lookup[short]
    return None


def main():
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)
    coord_map = {s["name"]: (s["lat"], s["lng"]) for s in data["schools"]}
    print(f"JSON 보정 좌표 로드: {len(coord_map)}개 학교")

    wb_src = xlrd.open_workbook(str(SRC_XLS))
    ws_src = wb_src.sheet_by_index(0)
    print(f"원본 .xls 로드: {ws_src.nrows}행 x {ws_src.ncols}열")

    wb_dst = Workbook()
    ws_dst = wb_dst.active
    ws_dst.title = ws_src.name

    # 헤더 복사
    for c in range(ws_src.ncols):
        ws_dst.cell(row=1, column=c + 1, value=ws_src.cell_value(0, c))

    updated = 0
    unchanged = 0
    missing = []

    for r in range(1, ws_src.nrows):
        name = (ws_src.cell_value(r, 0) or "").strip()
        for c in range(ws_src.ncols):
            val = ws_src.cell_value(r, c)
            # 위도(col 8), 경도(col 9)는 JSON 보정값으로 덮어씀
            if c == 8 or c == 9:
                continue
            ws_dst.cell(row=r + 1, column=c + 1, value=val)

        match = find_match(name, coord_map) if name else None
        if match:
            j_lat, j_lng = match
            try:
                e_lat = float(ws_src.cell_value(r, 8)) if ws_src.cell_value(r, 8) else 0
                e_lng = float(ws_src.cell_value(r, 9)) if ws_src.cell_value(r, 9) else 0
            except (ValueError, TypeError):
                e_lat, e_lng = 0, 0

            # 원본과 동일하게 텍스트 형태로 저장
            ws_dst.cell(row=r + 1, column=9, value=f"{j_lat:.7f}".rstrip("0").rstrip("."))
            ws_dst.cell(row=r + 1, column=10, value=f"{j_lng:.7f}".rstrip("0").rstrip("."))

            if abs(j_lat - e_lat) > 1e-6 or abs(j_lng - e_lng) > 1e-6:
                updated += 1
            else:
                unchanged += 1
        else:
            # JSON에서 못 찾은 경우 원본 값 그대로 유지
            ws_dst.cell(row=r + 1, column=9, value=ws_src.cell_value(r, 8))
            ws_dst.cell(row=r + 1, column=10, value=ws_src.cell_value(r, 9))
            if name:
                missing.append(name)

    wb_dst.save(str(DST_XLSX))
    print()
    print(f"=== 결과 ===")
    print(f"좌표 보정 반영: {updated}건")
    print(f"원본과 동일: {unchanged}건")
    print(f"JSON 미매칭 (원본 좌표 유지): {len(missing)}건")
    if missing:
        print(f"  미매칭 학교: {missing}")
    print(f"저장: {DST_XLSX}")


if __name__ == "__main__":
    main()
