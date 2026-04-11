"""
학교를 울산학교주소위도경도.xlsx와
울산학교개교일우편번호전화번호팩스번호홈페이지.xlsx에서 제거.

폐교 또는 잘못 입력된 학교를 정리할 때 사용.

사용 예:
  python scripts/remove_school.py --name "울산OO중학교"

동작:
  1. 두 엑셀 파일을 각각 .bak으로 백업
  2. 두 파일에서 해당 학교명이 포함된 행을 찾아 삭제
  3. 양쪽 다 찾지 못하면 경고만 출력하고 종료
"""

import argparse
import shutil
from pathlib import Path

import openpyxl

BASE = Path(__file__).parent.parent
XLSX_DIR = BASE / "data" / "xlsx"
LOC_XLSX = XLSX_DIR / "울산학교주소위도경도.xlsx"
CONTACT_XLSX = XLSX_DIR / "울산학교개교일우편번호전화번호팩스번호홈페이지.xlsx"


def remove_from(path, name_col, target):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows_to_delete = []
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, name_col).value
        if val and str(val).strip() == target:
            rows_to_delete.append(r)

    if not rows_to_delete:
        return 0

    # 뒤에서부터 삭제 (인덱스 밀림 방지)
    for r in reversed(rows_to_delete):
        ws.delete_rows(r, 1)
    wb.save(path)
    return len(rows_to_delete)


def main():
    p = argparse.ArgumentParser(description="학교를 엑셀 2개 파일에서 제거")
    p.add_argument("--name", required=True, help="제거할 학교명")
    args = p.parse_args()

    for path in (LOC_XLSX, CONTACT_XLSX):
        bak = path.with_suffix(path.suffix + ".bak")
        shutil.copy2(path, bak)
        print(f"백업: {bak.name}")

    loc_cnt = remove_from(LOC_XLSX, name_col=1, target=args.name)
    ct_cnt = remove_from(CONTACT_XLSX, name_col=2, target=args.name)

    print(f"[OK] 위치 엑셀에서 {loc_cnt}행 삭제")
    print(f"[OK] 연락처 엑셀에서 {ct_cnt}행 삭제")

    if loc_cnt == 0 and ct_cnt == 0:
        print(f"[WARN] '{args.name}'을(를) 양쪽 파일 어디에서도 찾지 못했습니다.")

    print()
    print("다음 단계: python scripts/import_excel.py 를 실행하여 JSON에 반영하세요.")


if __name__ == "__main__":
    main()
