"""
새 학교를 울산학교주소위도경도.xlsx와
울산학교개교일우편번호전화번호팩스번호홈페이지.xlsx 양쪽에 동시에 추가.

사용 예:
  python scripts/add_school.py \\
      --name "울산미래중학교" \\
      --type "중학교" \\
      --founded "2027-03-02" \\
      --founding-type "공립" \\
      --address-jibun "울산광역시 북구 ○○동 123" \\
      --address-road "울산광역시 북구 ○○로 45" \\
      --lat 35.5678 --lng 129.3456 \\
      --zipcode "44123" \\
      --phone "052-123-4567" \\
      --fax "052-123-4568" \\
      --homepage "http://example.es.kr/"

동작:
  1. 두 엑셀 파일을 각각 .bak으로 백업 (기존 .bak은 덮어씀)
  2. 학교명 중복 확인 (있으면 중단)
  3. 두 엑셀에 행 추가
  4. 저장
  5. 이후 python scripts/import_excel.py를 돌려서 JSON 반영 필요
"""

import argparse
import shutil
from pathlib import Path

import openpyxl

BASE = Path(__file__).parent.parent
LOC_XLSX = BASE / "울산학교주소위도경도.xlsx"
CONTACT_XLSX = BASE / "울산학교개교일우편번호전화번호팩스번호홈페이지.xlsx"


def find_existing(name):
    """두 파일에 이미 같은 학교명이 있는지 확인. 있으면 파일명 반환, 없으면 None."""
    for path, name_col in [(LOC_XLSX, 1), (CONTACT_XLSX, 2)]:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row_idx, name_col).value
            if val and str(val).strip() == name:
                return path.name
    return None


def add_to_location(args):
    wb = openpyxl.load_workbook(LOC_XLSX)
    ws = wb.active
    r = ws.max_row + 1
    ws.cell(r, 1, args.name)
    ws.cell(r, 2, args.type)
    ws.cell(r, 3, args.founded)
    ws.cell(r, 4, args.founding_type)
    ws.cell(r, 5, "본교")
    ws.cell(r, 6, "운영")
    ws.cell(r, 7, args.address_jibun)
    ws.cell(r, 8, args.address_road)
    ws.cell(r, 9, f"{args.lat}")
    ws.cell(r, 10, f"{args.lng}")
    wb.save(LOC_XLSX)
    return r


def add_to_contact(args):
    wb = openpyxl.load_workbook(CONTACT_XLSX)
    ws = wb.active
    r = ws.max_row + 1
    founded_nodash = args.founded.replace("-", "")
    ws.cell(r, 1, args.type)
    ws.cell(r, 2, args.name)
    ws.cell(r, 3, founded_nodash)
    ws.cell(r, 4, args.zipcode)
    ws.cell(r, 5, args.phone)
    ws.cell(r, 6, args.fax or "")
    ws.cell(r, 7, args.homepage or "")
    wb.save(CONTACT_XLSX)
    return r


def main():
    p = argparse.ArgumentParser(description="새 학교를 엑셀 2개 파일에 추가")
    p.add_argument("--name", required=True, help="학교명 (예: 울산미래중학교)")
    p.add_argument("--type", required=True, choices=["초등학교", "중학교", "고등학교"])
    p.add_argument("--founded", required=True, help="설립일자 YYYY-MM-DD")
    p.add_argument("--founding-type", required=True, help="공립/사립 등")
    p.add_argument("--address-jibun", required=True, help="지번 주소")
    p.add_argument("--address-road", required=True, help="도로명 주소")
    p.add_argument("--lat", required=True, type=float, help="위도 (WGS84)")
    p.add_argument("--lng", required=True, type=float, help="경도 (WGS84)")
    p.add_argument("--zipcode", required=True, help="우편번호")
    p.add_argument("--phone", required=True, help="전화번호")
    p.add_argument("--fax", default="", help="팩스번호 (선택)")
    p.add_argument("--homepage", default="", help="홈페이지 (선택)")
    args = p.parse_args()

    existing = find_existing(args.name)
    if existing:
        print(f"[ERROR] {args.name}은(는) 이미 {existing}에 존재합니다.")
        raise SystemExit(1)

    for path in (LOC_XLSX, CONTACT_XLSX):
        bak = path.with_suffix(path.suffix + ".bak")
        shutil.copy2(path, bak)
        print(f"백업: {bak.name}")

    loc_row = add_to_location(args)
    print(f"[OK] 위치 엑셀 {loc_row}행: {args.name}")

    ct_row = add_to_contact(args)
    print(f"[OK] 연락처 엑셀 {ct_row}행: {args.name}")

    print()
    print("다음 단계: python scripts/import_excel.py 를 실행하여 JSON에 반영하세요.")


if __name__ == "__main__":
    main()
